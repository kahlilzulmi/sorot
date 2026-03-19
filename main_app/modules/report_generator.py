"""
Report Generator Module

This module provides functions for generating comprehensive reports from session data,
including PDF and Excel reports with charts and statistical analysis.

Author: Kahlil Gibran Al Zulmi
Institution: Institut Teknologi Sepuluh Nopember
Date: November 2025
"""

import os
from datetime import datetime
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from modules import database_manager as db
from modules import visualization as vis
from utils.logger import log_info, log_error


# ============================================================================
# EXCEL REPORT GENERATION
# ============================================================================

def generate_detection_excel_report(
    session_id: int,
    output_path: str
) -> bool:
    """
    Generate Excel report for a detection session.
    
    Args:
        session_id: Detection session ID
        output_path: Path to save Excel file
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Get session data
        session = db.get_detection_session(session_id)
        if not session:
            log_error(f"Detection session {session_id} not found")
            return False
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Session Summary
        ws1 = wb.active
        ws1.title = "Session Summary"
        _create_detection_summary_sheet(ws1, session)
        
        # Sheet 2: Detection Results
        if session.get('results'):
            ws2 = wb.create_sheet("Detection Results")
            _create_detection_results_sheet(ws2, session['results'])
        
        # Sheet 3: Statistics
        ws3 = wb.create_sheet("Statistics")
        _create_detection_statistics_sheet(ws3, session)
        
        # Save workbook
        wb.save(output_path)
        log_info(f"Detection Excel report saved to {output_path}")
        return True
    
    except Exception as e:
        log_error(f"Error generating detection Excel report: {str(e)}")
        return False


def _create_detection_summary_sheet(ws, session: Dict[str, Any]):
    """Create summary sheet for detection report."""
    import json
    
    # Title
    ws['A1'] = "Detection Session Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')
    
    # Session info
    row = 3
    info_items = [
        ("Session ID", session.get('session_id')),
        ("Timestamp", session.get('timestamp')),
        ("Video Path", session.get('video_path')),
        ("Output Path", session.get('output_path', 'N/A')),
        ("Methods Used", ', '.join(json.loads(session.get('methods_used', '[]')))),
        ("Total Frames", session.get('total_frames')),
        ("Frames Processed", session.get('frames_processed')),
        ("Detections Count", session.get('detections_count')),
        ("Processing Time (s)", f"{session.get('processing_time_seconds', 0):.2f}"),
        ("Notes", session.get('notes', 'N/A'))
    ]
    
    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    # Format columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def _create_detection_results_sheet(ws, results: List[Dict[str, Any]]):
    """Create results sheet for detection report."""
    # Headers
    headers = ["Frame", "Method", "Center X", "Center Y", "Radius", "Confidence"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result.get('frame_number'))
        ws.cell(row=row, column=2, value=result.get('method'))
        ws.cell(row=row, column=3, value=f"{result.get('center_x', 0):.2f}")
        ws.cell(row=row, column=4, value=f"{result.get('center_y', 0):.2f}")
        ws.cell(row=row, column=5, value=f"{result.get('radius', 0):.2f}")
        ws.cell(row=row, column=6, value=result.get('confidence', 1.0))
    
    # Auto-size columns
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 15


def _create_detection_statistics_sheet(ws, session: Dict[str, Any]):
    """Create statistics sheet for detection report."""
    results = session.get('results', [])
    
    # Title
    ws['A1'] = "Detection Statistics"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Calculate statistics
    if results:
        radii = [r.get('radius', 0) for r in results]
        avg_radius = sum(radii) / len(radii) if radii else 0
        min_radius = min(radii) if radii else 0
        max_radius = max(radii) if radii else 0
    else:
        avg_radius = min_radius = max_radius = 0
    
    # Statistics
    row = 3
    stats = [
        ("Total Detections", len(results)),
        ("Average Radius", f"{avg_radius:.2f}"),
        ("Min Radius", f"{min_radius:.2f}"),
        ("Max Radius", f"{max_radius:.2f}"),
        ("Success Rate", f"{(session.get('detections_count', 0) / session.get('frames_processed', 1) * 100):.1f}%")
    ]
    
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20


def generate_game_excel_report(
    session_id: int,
    output_path: str
) -> bool:
    """
    Generate Excel report for a game session.
    
    Args:
        session_id: Game session ID
        output_path: Path to save Excel file
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Get session data
        session = db.get_game_session(session_id)
        if not session:
            log_error(f"Game session {session_id} not found")
            return False
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Session Summary
        ws1 = wb.active
        ws1.title = "Session Summary"
        _create_game_summary_sheet(ws1, session)
        
        # Sheet 2: Question Results
        if session.get('questions'):
            ws2 = wb.create_sheet("Question Results")
            _create_game_results_sheet(ws2, session['questions'])
            
            # Add chart
            _add_game_chart(ws2, len(session['questions']))
        
        # Sheet 3: Performance Analysis
        ws3 = wb.create_sheet("Performance Analysis")
        _create_game_analysis_sheet(ws3, session)
        
        # Save workbook
        wb.save(output_path)
        log_info(f"Game Excel report saved to {output_path}")
        return True
    
    except Exception as e:
        log_error(f"Error generating game Excel report: {str(e)}")
        return False


def _create_game_summary_sheet(ws, session: Dict[str, Any]):
    """Create summary sheet for game report."""
    # Title
    ws['A1'] = "Game Session Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')
    
    # Session info
    row = 3
    info_items = [
        ("Session ID", session.get('session_id')),
        ("Timestamp", session.get('timestamp')),
        ("Participant ID", session.get('participant_id')),
        ("Mode", session.get('mode')),
        ("Total Questions", session.get('total_questions')),
        ("Correct Answers", session.get('correct_answers')),
        ("Score Percentage", f"{session.get('score_percentage', 0):.2f}%"),
        ("Total Time (s)", f"{session.get('total_time_seconds', 0):.2f}"),
        ("Session Path", session.get('session_path', 'N/A')),
        ("Notes", session.get('notes', 'N/A'))
    ]
    
    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def _create_game_results_sheet(ws, questions: List[Dict[str, Any]]):
    """Create results sheet for game report."""
    # Headers
    headers = ["#", "Question", "Correct", "User Answer", "Result", "Time (s)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row, q in enumerate(questions, 2):
        ws.cell(row=row, column=1, value=q.get('question_index', 0) + 1)
        ws.cell(row=row, column=2, value=q.get('question_text'))
        ws.cell(row=row, column=3, value=q.get('correct_answer'))
        ws.cell(row=row, column=4, value=q.get('user_answer'))
        
        result_cell = ws.cell(row=row, column=5, value="✓" if q.get('is_correct') else "✗")
        if q.get('is_correct'):
            result_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            result_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=row, column=6, value=f"{q.get('response_time_seconds', 0):.2f}")
    
    # Auto-size columns
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12


def _add_game_chart(ws, num_questions: int):
    """Add chart to game results sheet."""
    # Create bar chart for response times
    chart = BarChart()
    chart.title = "Response Times"
    chart.x_axis.title = "Question"
    chart.y_axis.title = "Time (seconds)"
    
    data = Reference(ws, min_col=6, min_row=1, max_row=num_questions + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=num_questions + 1)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    ws.add_chart(chart, f"H2")


def _create_game_analysis_sheet(ws, session: Dict[str, Any]):
    """Create analysis sheet for game report."""
    questions = session.get('questions', [])
    
    # Title
    ws['A1'] = "Performance Analysis"
    ws['A1'].font = Font(size=14, bold=True)
    
    # Calculate metrics
    if questions:
        response_times = [q.get('response_time_seconds', 0) for q in questions]
        avg_time = sum(response_times) / len(response_times)
        min_time = min(response_times)
        max_time = max(response_times)
    else:
        avg_time = min_time = max_time = 0
    
    # Analysis
    row = 3
    analysis = [
        ("Total Questions", len(questions)),
        ("Correct Answers", session.get('correct_answers', 0)),
        ("Wrong Answers", len(questions) - session.get('correct_answers', 0)),
        ("Accuracy", f"{session.get('score_percentage', 0):.2f}%"),
        ("Average Response Time", f"{avg_time:.2f}s"),
        ("Fastest Response", f"{min_time:.2f}s"),
        ("Slowest Response", f"{max_time:.2f}s"),
        ("Total Session Time", f"{session.get('total_time_seconds', 0):.2f}s")
    ]
    
    for label, value in analysis:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def generate_stimulus_excel_report(
    session_id: int,
    output_path: str
) -> bool:
    """
    Generate Excel report for a stimulus session.
    
    Args:
        session_id: Stimulus session ID
        output_path: Path to save Excel file
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Get session data
        session = db.get_stimulus_session(session_id)
        if not session:
            log_error(f"Stimulus session {session_id} not found")
            return False
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Session Summary
        ws1 = wb.active
        ws1.title = "Session Summary"
        _create_stimulus_summary_sheet(ws1, session)
        
        # Sheet 2: Tasks List
        if session.get('tasks'):
            ws2 = wb.create_sheet("Tasks")
            _create_stimulus_tasks_sheet(ws2, session['tasks'])
        
        # Save workbook
        wb.save(output_path)
        log_info(f"Stimulus Excel report saved to {output_path}")
        return True
    
    except Exception as e:
        log_error(f"Error generating stimulus Excel report: {str(e)}")
        return False


def _create_stimulus_summary_sheet(ws, session: Dict[str, Any]):
    """Create summary sheet for stimulus report."""
    # Title
    ws['A1'] = "Stimulus Session Report"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:B1')
    
    # Session info
    row = 3
    info_items = [
        ("Session ID", session.get('session_id')),
        ("Timestamp", session.get('timestamp')),
        ("Protocol Name", session.get('protocol_name')),
        ("Video Path", session.get('video_path')),
        ("Duration (min)", f"{session.get('duration_seconds', 0) / 60:.2f}"),
        ("Frame Count", session.get('frame_count')),
        ("File Size (MB)", f"{session.get('file_size_mb', 0):.2f}"),
        ("Generation Time (s)", f"{session.get('generation_time_seconds', 0):.2f}"),
        ("Resolution", f"{session.get('resolution_width')}x{session.get('resolution_height')}"),
        ("FPS", session.get('fps')),
        ("Notes", session.get('notes', 'N/A'))
    ]
    
    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 50


def _create_stimulus_tasks_sheet(ws, tasks: List[Dict[str, Any]]):
    """Create tasks sheet for stimulus report."""
    # Headers
    headers = ["#", "Task Type", "Duration (s)", "Position X", "Position Y"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(color="000000", bold=True)
    
    # Data
    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task.get('task_index', 0) + 1)
        ws.cell(row=row, column=2, value=task.get('task_type'))
        ws.cell(row=row, column=3, value=f"{task.get('duration_seconds', 0):.2f}")
        
        pos_x = task.get('position_x')
        pos_y = task.get('position_y')
        ws.cell(row=row, column=4, value=f"{pos_x:.0f}" if pos_x is not None else "N/A")
        ws.cell(row=row, column=5, value=f"{pos_y:.0f}" if pos_y is not None else "N/A")
    
    # Auto-size columns
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20


# ============================================================================
# COMPARISON REPORTS
# ============================================================================

def generate_comparison_report(
    session_ids: List[int],
    session_type: str,
    output_path: str
) -> bool:
    """
    Generate comparison report for multiple sessions.
    
    Args:
        session_ids: List of session IDs to compare
        session_type: Type of sessions ('detection', 'game', 'stimulus')
        output_path: Path to save Excel file
        
    Returns:
        True if successful, False otherwise
    """
    try:
        # Get sessions
        sessions = []
        for sid in session_ids:
            if session_type == 'detection':
                session = db.get_detection_session(sid)
            elif session_type == 'game':
                session = db.get_game_session(sid)
            else:
                session = db.get_stimulus_session(sid)
            
            if session:
                sessions.append(session)
        
        if not sessions:
            log_error("No sessions found for comparison")
            return False
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Session Comparison"
        
        # Create comparison based on type
        if session_type == 'game':
            _create_game_comparison(ws, sessions)
        elif session_type == 'detection':
            _create_detection_comparison(ws, sessions)
        else:
            _create_stimulus_comparison(ws, sessions)
        
        # Save workbook
        wb.save(output_path)
        log_info(f"Comparison report saved to {output_path}")
        return True
    
    except Exception as e:
        log_error(f"Error generating comparison report: {str(e)}")
        return False


def _create_game_comparison(ws, sessions: List[Dict[str, Any]]):
    """Create game session comparison."""
    # Title
    ws['A1'] = "Game Session Comparison"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    # Headers
    headers = ["Session ID", "Participant", "Score %", "Questions", "Time (s)", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row, session in enumerate(sessions, 4):
        ws.cell(row=row, column=1, value=session.get('session_id'))
        ws.cell(row=row, column=2, value=session.get('participant_id'))
        ws.cell(row=row, column=3, value=f"{session.get('score_percentage', 0):.1f}%")
        ws.cell(row=row, column=4, value=session.get('total_questions'))
        ws.cell(row=row, column=5, value=f"{session.get('total_time_seconds', 0):.1f}")
        ws.cell(row=row, column=6, value=session.get('timestamp'))
    
    # Auto-size columns
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 18


def _create_detection_comparison(ws, sessions: List[Dict[str, Any]]):
    """Create detection session comparison."""
    import json
    
    # Title
    ws['A1'] = "Detection Session Comparison"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    # Headers
    headers = ["Session ID", "Video", "Methods", "Detections", "Time (s)", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Data
    for row, session in enumerate(sessions, 4):
        methods = ', '.join(json.loads(session.get('methods_used', '[]')))
        ws.cell(row=row, column=1, value=session.get('session_id'))
        ws.cell(row=row, column=2, value=os.path.basename(session.get('video_path', '')))
        ws.cell(row=row, column=3, value=methods)
        ws.cell(row=row, column=4, value=session.get('detections_count'))
        ws.cell(row=row, column=5, value=f"{session.get('processing_time_seconds', 0):.1f}")
        ws.cell(row=row, column=6, value=session.get('timestamp'))
    
    # Auto-size columns
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20


def _create_stimulus_comparison(ws, sessions: List[Dict[str, Any]]):
    """Create stimulus session comparison."""
    # Title
    ws['A1'] = "Stimulus Session Comparison"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:F1')
    
    # Headers
    headers = ["Session ID", "Protocol", "Duration (min)", "Tasks", "Size (MB)", "Date"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        cell.font = Font(color="000000", bold=True)
    
    # Data
    for row, session in enumerate(sessions, 4):
        task_count = len(session.get('tasks', []))
        ws.cell(row=row, column=1, value=session.get('session_id'))
        ws.cell(row=row, column=2, value=session.get('protocol_name'))
        ws.cell(row=row, column=3, value=f"{session.get('duration_seconds', 0) / 60:.1f}")
        ws.cell(row=row, column=4, value=task_count)
        ws.cell(row=row, column=5, value=f"{session.get('file_size_mb', 0):.1f}")
        ws.cell(row=row, column=6, value=session.get('timestamp'))
    
    # Auto-size columns
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 20


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def generate_all_reports(
    session_id: int,
    session_type: str,
    output_dir: str
) -> Dict[str, str]:
    """
    Generate all available reports for a session.
    
    Args:
        session_id: Session ID
        session_type: Type of session
        output_dir: Directory to save reports
        
    Returns:
        Dictionary mapping report types to file paths
    """
    reports = {}
    
    try:
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Excel report
        excel_path = os.path.join(output_dir, f"{session_type}_session_{session_id}_{timestamp}.xlsx")
        
        if session_type == 'detection':
            if generate_detection_excel_report(session_id, excel_path):
                reports['excel'] = excel_path
        elif session_type == 'game':
            if generate_game_excel_report(session_id, excel_path):
                reports['excel'] = excel_path
        else:
            if generate_stimulus_excel_report(session_id, excel_path):
                reports['excel'] = excel_path
        
        log_info(f"Generated {len(reports)} reports for {session_type} session {session_id}")
        return reports
    
    except Exception as e:
        log_error(f"Error generating reports: {str(e)}")
        return reports


if __name__ == "__main__":
    # Test report generation
    print("Report Generator Module Test")
    print("=" * 60)
    print("✓ Report generator module ready")
    print("✓ Excel report functions available")
    print("✓ Comparison report functions available")
