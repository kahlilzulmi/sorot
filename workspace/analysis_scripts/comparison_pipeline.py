# ==============================================================================
# IMPORTS
# ==============================================================================
import cv2
import numpy as np
import pandas as pd
import os
import json
import traceback
import datetime
from tqdm import tqdm
from itertools import combinations
import scipy.stats
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend for memory efficiency
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

print("=" * 80)
print("Eye Gaze Detection Algorithm Comparison System")
print("=" * 80)
print("All imports loaded successfully.")

# ==============================================================================
# CONSTANTS AND CONFIGURATION
# ==============================================================================
# Video properties (must match ground truth)
WIDTH = 1920
HEIGHT = 1080
FPS = 60

# Detection parameters
CENTER_FALLBACK = (960, 540)  # Center screen fallback for missing detections
MAX_OFFSET = 300  # Maximum frame offset for cross-correlation (±5 seconds)

# Threshold values
DETECTION_THRESHOLD = 95.0  # Minimum detection rate percentage
DISTANCE_THRESHOLD = 75.0  # Maximum acceptable euclidean distance in pixels (gaze area radius)

# File paths
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
WORKSPACE_ROOT = os.path.dirname(WORKSPACE_DIR)  # workspace/
PROJECT_ROOT = os.path.dirname(WORKSPACE_ROOT)   # tugasakhir/

OUTPUT_BASE_DIR = os.path.join(WORKSPACE_DIR, "output")
DETECTION_RESULTS_DIR = os.path.join(OUTPUT_BASE_DIR, "detection_results")
PLOTS_DIR = os.path.join(OUTPUT_BASE_DIR, "plots")
REPORTS_DIR = os.path.join(OUTPUT_BASE_DIR, "reports")
VIDEO_OVERLAYS_DIR = os.path.join(OUTPUT_BASE_DIR, "video_overlays")
CHECKPOINTS_DIR = os.path.join(OUTPUT_BASE_DIR, "checkpoints")
LOGS_DIR = os.path.join(OUTPUT_BASE_DIR, "logs")

VIDEO_DIR = os.path.join(WORKSPACE_ROOT, "eye_gaze", "rapi")
GT_CSV_PATH = os.path.join(PROJECT_ROOT, "Archived", "stimulus_ground_truth_trimmed.csv")

# Method and variant names
METHODS = ["color", "contour", "blob", "hough", "kalman", "kalman2"]
VARIANTS = ["varian1", "varian2", "varian3"]

print(f"Configuration loaded:")
print(f"  - Video resolution: {WIDTH}x{HEIGHT} @ {FPS} FPS")
print(f"  - Detection threshold: {DETECTION_THRESHOLD}%")
print(f"  - Distance threshold: {DISTANCE_THRESHOLD} pixels")
print(f"  - Output directory: {OUTPUT_BASE_DIR}")

# ==============================================================================
# CHECKPOINT MANAGER CLASS
# ==============================================================================
class CheckpointManager:
    """Manages checkpoint state for resumable processing."""
    
    def __init__(self, checkpoint_file=None):
        if checkpoint_file is None:
            checkpoint_file = os.path.join(CHECKPOINTS_DIR, "progress.json")
        self.checkpoint_file = checkpoint_file
        self.data = self.load_checkpoint()
    
    def load_checkpoint(self):
        """Load checkpoint from JSON file."""
        if os.path.exists(self.checkpoint_file):
            try:
                with open(self.checkpoint_file, 'r') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Warning: Could not load checkpoint: {e}")
                return {"completed": [], "timestamp": None}
        return {"completed": [], "timestamp": None}
    
    def save_progress(self, method, variant):
        """Mark a method-variant combination as completed."""
        key = f"{method}_{variant}"
        if key not in self.data["completed"]:
            self.data["completed"].append(key)
        self.data["timestamp"] = datetime.datetime.now().isoformat()
        
        os.makedirs(os.path.dirname(self.checkpoint_file), exist_ok=True)
        with open(self.checkpoint_file, 'w') as f:
            json.dump(self.data, f, indent=2)
    
    def is_completed(self, method, variant):
        """Check if a method-variant combination is already completed."""
        key = f"{method}_{variant}"
        return key in self.data["completed"]
    
    def clear(self):
        """Clear all checkpoint data."""
        self.data = {"completed": [], "timestamp": None}
        if os.path.exists(self.checkpoint_file):
            os.remove(self.checkpoint_file)
    
    def get_progress(self):
        """Get completion progress as tuple (completed, total)."""
        total = len(METHODS) * len(VARIANTS)
        completed = len(self.data["completed"])
        return completed, total


# ==============================================================================
# ERROR LOGGER CLASS
# ==============================================================================
class ErrorLogger:
    """Logs detailed error information to file."""
    
    def __init__(self, log_dir=None):
        if log_dir is None:
            log_dir = LOGS_DIR
        self.log_dir = log_dir
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_file = os.path.join(log_dir, f"error_log_{timestamp}.txt")
        os.makedirs(log_dir, exist_ok=True)
        
        # Create initial log entry
        with open(self.log_file, 'w') as f:
            f.write("=" * 80 + "\n")
            f.write("ERROR LOG - Eye Gaze Detection Comparison Pipeline\n")
            f.write("=" * 80 + "\n")
            f.write(f"Started: {datetime.datetime.now().isoformat()}\n")
            f.write("\n")
    
    def log_error(self, method, variant, frame_num, exception):
        """Log an error with full context and traceback."""
        with open(self.log_file, 'a') as f:
            f.write("-" * 80 + "\n")
            f.write(f"ERROR at {datetime.datetime.now().isoformat()}\n")
            f.write(f"Method: {method}\n")
            f.write(f"Variant: {variant}\n")
            f.write(f"Frame: {frame_num}\n")
            f.write(f"Exception Type: {type(exception).__name__}\n")
            f.write(f"Exception Message: {str(exception)}\n")
            f.write("\nFull Traceback:\n")
            f.write(traceback.format_exc())
            f.write("\n")
        
        print(f"\n{'!'*80}")
        print(f"ERROR: {method}_{variant} failed at frame {frame_num}")
        print(f"Exception: {type(exception).__name__}: {str(exception)}")
        print(f"Details logged to: {self.log_file}")
        print(f"{'!'*80}\n")

print("Checkpoint and Error logging systems initialized.")

# ==============================================================================
# UTILITY FUNCTIONS
# ==============================================================================

def validate_video_properties():
    """Validate that all 3 video variants have consistent properties."""
    print("\n" + "=" * 80)
    print("VALIDATING VIDEO PROPERTIES")
    print("=" * 80)
    
    properties = {}
    log_path = os.path.join(LOGS_DIR, "video_validation.txt")
    os.makedirs(LOGS_DIR, exist_ok=True)
    
    with open(log_path, 'w') as log_file:
        log_file.write("Video Properties Validation\n")
        log_file.write(f"Timestamp: {datetime.datetime.now().isoformat()}\n")
        log_file.write("=" * 80 + "\n\n")
        
        for variant in VARIANTS:
            video_path = os.path.join(VIDEO_DIR, f"{variant}.mp4")
            
            if not os.path.exists(video_path):
                error_msg = f"ERROR: Video file not found: {video_path}"
                print(error_msg)
                log_file.write(error_msg + "\n")
                continue
            
            cap = cv2.VideoCapture(video_path)
            if not cap.isOpened():
                error_msg = f"ERROR: Cannot open video: {video_path}"
                print(error_msg)
                log_file.write(error_msg + "\n")
                cap.release()
                continue
            
            width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
            height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            cap.release()
            
            properties[variant] = {
                'width': width,
                'height': height,
                'fps': fps,
                'frame_count': frame_count,
                'path': video_path
            }
            
            info = f"{variant}: {width}x{height} @ {fps:.2f} FPS, {frame_count} frames"
            print(f"   {info}")
            log_file.write(info + "\n")
        
        # Check consistency
        if len(properties) == len(VARIANTS):
            widths = [p['width'] for p in properties.values()]
            heights = [p['height'] for p in properties.values()]
            fps_values = [p['fps'] for p in properties.values()]
            
            if len(set(widths)) > 1 or len(set(heights)) > 1 or len(set(fps_values)) > 1:
                warning = "\nWARNING: Videos have inconsistent properties!"
                print(warning)
                log_file.write(warning + "\n")
            else:
                success = "\n All videos have consistent properties."
                print(success)
                log_file.write(success + "\n")
        
        log_file.write(f"\nValidation log saved to: {log_path}\n")
    
    print(f"Validation log saved to: {log_path}")
    return properties


# ==============================================================================
# DETECTION ALGORITHM FUNCTIONS
# ==============================================================================

def detect_color_frame(frame, state):
    """Detect using HSV color-based method."""
    # HSV parameters (from detect_color.py)
    h_min, h_max = 80, 100
    s_min, s_max = 50, 255
    v_min, v_max = 200, 255
    
    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
    lower_bound = np.array([h_min, s_min, v_min])
    upper_bound = np.array([h_max, s_max, v_max])
    mask = cv2.inRange(hsv, lower_bound, upper_bound)
    
    kernel = np.ones((3,3), np.uint8)
    mask = cv2.erode(mask, kernel, iterations=1)
    mask = cv2.dilate(mask, kernel, iterations=2)
    
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if contours:
        largest_contour = max(contours, key=cv2.contourArea)
        if cv2.contourArea(largest_contour) > 10:
            M = cv2.moments(largest_contour)
            if M["m00"] != 0:
                x = int(M["m10"] / M["m00"])
                y = int(M["m01"] / M["m00"])
                return (x, y)
    
    return None


def detect_contour_frame(frame, state):
    """Detect using adaptive contour-based method."""
    BBOX_SIZE = state.get('bbox_size', 400)
    MIN_AREA = 50
    
    hsv = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
    
    # Determine ROI
    if 'roi' in state and state['roi'] is not None:
        x1, y1, x2, y2 = state['roi']
        # Validate ROI coordinates
        if x1 >= x2 or y1 >= y2 or x1 < 0 or y1 < 0:
            # Invalid ROI, reset and use full frame
            state['roi'] = None
            roi = frame
            roi_origin = (0, 0)
        else:
            roi = frame[y1:y2, x1:x2]
            # Check if ROI is empty after slicing
            if roi.size == 0:
                state['roi'] = None
                roi = frame
                roi_origin = (0, 0)
            else:
                roi_origin = (x1, y1)
    else:
        roi = frame
        roi_origin = (0, 0)
    
    # Try sensitive mode first
    hsv_roi = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
    lower_cyan = np.array([80, 40, 40])
    upper_cyan = np.array([100, 255, 255])
    mask = cv2.inRange(hsv_roi, lower_cyan, upper_cyan)
    
    kernel_small = np.ones((3,3), np.uint8)
    mask = cv2.dilate(mask, kernel_small, iterations=1)
    
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if contours:
        largest_contour = max(contours, key=cv2.contourArea)
        if cv2.contourArea(largest_contour) >= MIN_AREA and len(largest_contour) >= 5:
            ellipse = cv2.fitEllipse(largest_contour)
            cx, cy = int(ellipse[0][0]) + roi_origin[0], int(ellipse[0][1]) + roi_origin[1]
            
            # Update ROI for next frame
            height, width = frame.shape[:2]
            x1 = max(0, cx - BBOX_SIZE // 2)
            y1 = max(0, cy - BBOX_SIZE // 2)
            x2 = min(width, cx + BBOX_SIZE // 2)
            y2 = min(height, cy + BBOX_SIZE // 2)
            state['roi'] = (x1, y1, x2, y2)
            
            return (cx, cy)
    
    # Fallback: reset ROI
    state['roi'] = None
    return None


def detect_blob_frame(frame, state):
    """Detect using SimpleBlobDetector."""
    if 'detector' not in state:
        params = cv2.SimpleBlobDetector_Params()
        params.minThreshold = 10
        params.maxThreshold = 255
        params.filterByColor = True
        params.blobColor = 255
        params.filterByArea = True
        params.minArea = 30
        params.maxArea = 5000
        params.filterByCircularity = False
        params.filterByConvexity = False
        params.filterByInertia = False
        state['detector'] = cv2.SimpleBlobDetector_create(params)
    
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    keypoints = state['detector'].detect(gray)
    
    if keypoints:
        kp = keypoints[0]
        x, y = int(kp.pt[0]), int(kp.pt[1])
        
        # Refine center
        radius = int(kp.size / 2) + 5
        mask = np.zeros_like(gray)
        cv2.circle(mask, (x, y), radius, (255), -1)
        masked_gray = cv2.bitwise_and(gray, gray, mask=mask)
        _, binary_blob = cv2.threshold(masked_gray, 180, 255, cv2.THRESH_BINARY)
        moments = cv2.moments(binary_blob)
        if moments['m00'] != 0:
            cx = int(moments['m10'] / moments['m00'])
            cy = int(moments['m01'] / moments['m00'])
            return (cx, cy)
        return (x, y)
    
    return None


def detect_hough_frame(frame, state):
    """Detect using Hough Circle Transform."""
    HOUGH_PARAM1 = 50
    HOUGH_PARAM2 = 13
    MIN_RADIUS = 70
    MAX_RADIUS = 75
    
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    gray = cv2.medianBlur(gray, 5)
    
    circles = cv2.HoughCircles(
        gray,
        cv2.HOUGH_GRADIENT,
        dp=1,
        minDist=gray.shape[0],
        param1=HOUGH_PARAM1,
        param2=HOUGH_PARAM2,
        minRadius=MIN_RADIUS,
        maxRadius=MAX_RADIUS
    )
    
    if circles is not None:
        circles = np.uint16(np.around(circles))
        circle = circles[0][0]
        return (int(circle[0]), int(circle[1]))
    
    return None


def detect_kalman_frame(frame, state):
    """Detect using Hough Circle + Kalman Filter."""
    # Initialize Kalman filter on first call
    if 'kalman' not in state:
        kalman = cv2.KalmanFilter(4, 2)
        kalman.measurementMatrix = np.array([[1, 0, 0, 0], [0, 1, 0, 0]], np.float32)
        kalman.transitionMatrix = np.array([[1, 0, 1, 0], [0, 1, 0, 1], [0, 0, 1, 0], [0, 0, 0, 1]], np.float32)
        kalman.processNoiseCov = np.array([[1, 0, 0, 0], [0, 1, 0, 0], [0, 0, 1, 0], [0, 0, 0, 1]], np.float32) * 0.1
        kalman.measurementNoiseCov = np.array([[1, 0], [0, 1]], np.float32) * 1.0
        state['kalman'] = kalman
        state['kalman_initialized'] = False
    
    kalman = state['kalman']
    
    # Predict
    prediction = kalman.predict()
    
    # Detect using Hough
    detected = detect_hough_frame(frame, state)
    
    if detected is not None:
        measurement = np.array([[np.float32(detected[0])], [np.float32(detected[1])]])
        if not state['kalman_initialized']:
            kalman.statePost = np.array([measurement[0,0], measurement[1,0], 0, 0], np.float32)
            state['kalman_initialized'] = True
        else:
            kalman.correct(measurement)
    
    # Return Kalman state
    final_x = int(kalman.statePost[0])
    final_y = int(kalman.statePost[1])
    return (final_x, final_y)


def detect_kalman2_frame(frame, state):
    """Detect using Hough Circle + Smoother Kalman Filter."""
    # Initialize Kalman filter on first call
    if 'kalman' not in state:
        kalman = cv2.KalmanFilter(4, 2)
        kalman.measurementMatrix = np.array([[1, 0, 0, 0], [0, 1, 0, 0]], np.float32)
        kalman.transitionMatrix = np.array([[1, 0, 1, 0], [0, 1, 0, 1], [0, 0, 1, 0], [0, 0, 0, 1]], np.float32)
        kalman.processNoiseCov = np.array([[1, 0, 0, 0], [0, 1, 0, 0], [0, 0, 1, 0], [0, 0, 0, 1]], np.float32) * 0.1
        kalman.measurementNoiseCov = np.array([[1, 0], [0, 1]], np.float32) * 2.0  # Smoother
        state['kalman'] = kalman
        state['kalman_initialized'] = False
    
    kalman = state['kalman']
    
    # Predict
    prediction = kalman.predict()
    
    # Detect using Hough
    detected = detect_hough_frame(frame, state)
    
    if detected is not None:
        measurement = np.array([[np.float32(detected[0])], [np.float32(detected[1])]])
        if not state['kalman_initialized']:
            kalman.statePost = np.array([measurement[0,0], measurement[1,0], 0, 0], np.float32)
            state['kalman_initialized'] = True
        else:
            kalman.correct(measurement)
    
    # Return Kalman state
    final_x = int(kalman.statePost[0])
    final_y = int(kalman.statePost[1])
    return (final_x, final_y)


# Method dispatcher
DETECTION_FUNCTIONS = {
    'color': detect_color_frame,
    'contour': detect_contour_frame,
    'blob': detect_blob_frame,
    'hough': detect_hough_frame,
    'kalman': detect_kalman_frame,
    'kalman2': detect_kalman2_frame
}

print("Detection algorithms loaded:")
for method in METHODS:
    print(f"  - {method}")


# ==============================================================================
# MAIN DETECTION PROCESSING
# ==============================================================================

def process_single_video(video_path, method, output_csv):
    """
    Process a single video with a specific detection method.
    
    Args:
        video_path: Path to input video file
        method: Detection method name ('color', 'contour', etc.)
        output_csv: Path to save detection results CSV
    
    Returns:
        DataFrame with detection results
    """
    detection_func = DETECTION_FUNCTIONS[method]
    
    # Open video
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise ValueError(f"Cannot open video: {video_path}")
    
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    
    # Initialize state
    state = {'last_pos': CENTER_FALLBACK, 'bbox_size': 400}
    
    # Process frames
    results = []
    fallback_count = 0
    frame_idx = 0
    
    pbar = tqdm(total=total_frames, desc=f"  Processing frames", unit="frame", leave=False)
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        
        # Detect
        detected = detection_func(frame, state)
        
        # Handle missing detection
        is_interpolated = False
        if detected is None:
            detected = state['last_pos']
            is_interpolated = True
            fallback_count += 1
        else:
            state['last_pos'] = detected
        
        results.append({
            'frame': frame_idx,
            'x': detected[0],
            'y': detected[1],
            'is_interpolated': is_interpolated
        })
        
        frame_idx += 1
        pbar.update(1)
    
    pbar.close()
    cap.release()
    
    # Save CSV immediately
    df = pd.DataFrame(results)
    os.makedirs(os.path.dirname(output_csv), exist_ok=True)
    df.to_csv(output_csv, index=False)
    
    fallback_pct = (fallback_count / len(results)) * 100 if len(results) > 0 else 0
    print(f"    Fallbacks: {fallback_count}/{len(results)} ({fallback_pct:.2f}%)")
    
    return df


def run_detection_phase(checkpoint_manager, error_logger):
    """
    Run detection on all method-variant combinations with checkpoint support.
    
    Args:
        checkpoint_manager: CheckpointManager instance
        error_logger: ErrorLogger instance
    """
    print("\n" + "=" * 80)
    print("DETECTION PHASE")
    print("=" * 80)
    
    total_combinations = len(METHODS) * len(VARIANTS)
    completed, _ = checkpoint_manager.get_progress()
    
    # Overall progress bar
    overall_pbar = tqdm(total=total_combinations, initial=completed, desc="Overall Progress", unit="combo")
    
    for method in METHODS:
        for variant in VARIANTS:
            # Check checkpoint
            if checkpoint_manager.is_completed(method, variant):
                print(f"✓ Skipping {method}_{variant} (already completed)")
                continue
            
            combo_name = f"{method}_{variant}"
            print(f"\n[{method.upper()}] Processing {variant}...")
            
            try:
                video_path = os.path.join(VIDEO_DIR, f"{variant}.mp4")
                output_csv = os.path.join(DETECTION_RESULTS_DIR, f"{combo_name}.csv")
                
                # Process video
                df = process_single_video(video_path, method, output_csv)
                
                # Mark as completed
                checkpoint_manager.save_progress(method, variant)
                print(f"  ✓ Saved to: {output_csv}")
                
            except Exception as e:
                error_logger.log_error(method, variant, -1, e)
                print(f"\n{'!' * 80}")
                print(f"FATAL ERROR: Detection failed for {combo_name}")
                print(f"Pipeline stopped. Check error log for details.")
                print(f"{'!' * 80}\n")
                overall_pbar.close()
                raise
            
            overall_pbar.update(1)
    
    overall_pbar.close()
    print("\n" + "=" * 80)
    print("DETECTION PHASE COMPLETED")
    print("=" * 80)


# ==============================================================================
# SYNC AND METRICS CALCULATION
# ==============================================================================

def calculate_offset_limited(detected_csv, gt_csv):
    """
    Calculate optimal frame offset using cross-correlation with limited range.
    
    Args:
        detected_csv: Path to detected gaze CSV
        gt_csv: Path to ground truth CSV
    
    Returns:
        dict with offset_frames, delay_ms, correlation_coef
    """
    # Load data
    detected_df = pd.read_csv(detected_csv)
    gt_df = pd.read_csv(gt_csv)
    
    # Filter GT for valid frames
    gt_valid = gt_df[gt_df['gt_x_px'].notna()].copy()
    
    if len(gt_valid) == 0:
        return {'offset_frames': 0, 'delay_ms': 0.0, 'correlation_coef': 0.0}
    
    # Extract x trajectories
    gt_x = gt_valid['gt_x_px'].values
    
    # Match detected frames to GT frame range
    min_frame = gt_valid['frame'].min()
    max_frame = gt_valid['frame'].max()
    detected_filtered = detected_df[(detected_df['frame'] >= min_frame - MAX_OFFSET) & 
                                     (detected_df['frame'] <= max_frame + MAX_OFFSET)]
    
    if len(detected_filtered) < len(gt_x):
        return {'offset_frames': 0, 'delay_ms': 0.0, 'correlation_coef': 0.0}
    
    det_x = detected_filtered['x'].values
    
    # Normalize trajectories
    if gt_x.std() > 0:
        gt_x_norm = (gt_x - gt_x.min()) / (gt_x.max() - gt_x.min())
    else:
        gt_x_norm = gt_x
    
    if det_x.std() > 0:
        det_x_norm = (det_x - det_x.min()) / (det_x.max() - det_x.min())
    else:
        det_x_norm = det_x
    
    # Cross-correlation
    correlation = np.correlate(det_x_norm[:len(gt_x_norm)*2], gt_x_norm, mode='valid')
    
    if len(correlation) == 0:
        return {'offset_frames': 0, 'delay_ms': 0.0, 'correlation_coef': 0.0}
    
    # Find peak within allowed range
    peak_idx = np.argmax(correlation)
    offset_frames = peak_idx - len(gt_x_norm) // 2
    
    # Cap offset
    offset_frames = np.clip(offset_frames, -MAX_OFFSET, MAX_OFFSET)
    delay_ms = (offset_frames / FPS) * 1000
    correlation_coef = float(correlation[peak_idx]) if len(correlation) > 0 else 0.0
    
    return {
        'offset_frames': int(offset_frames),
        'delay_ms': float(delay_ms),
        'correlation_coef': float(correlation_coef)
    }


def calculate_task_specific_accuracy(merged_df):
    """
    Calculate accuracy breakdown by task type (movement pattern).
    
    Args:
        merged_df: Merged dataframe with 'phase' and 'within_radius' columns
    
    Returns:
        dict with accuracy per task type
    """
    task_breakdown = {}
    
    # Define task type patterns
    task_types = {
        'Horizontal Smooth Pursuit': ['HORIZONTAL'],
        'Vertical Smooth Pursuit': ['VERTICAL'],
        'Circular Smooth Pursuit': ['CIRCULAR'],
        'Structured Saccades': ['SACCADES', 'STRUCTURED'],
        'Random Saccades': ['SACCADES', 'RANDOM']
    }
    
    for task_name, patterns in task_types.items():
        # Filter frames matching all patterns
        mask = merged_df['phase'].str.contains(patterns[0], case=False, na=False)
        for pattern in patterns[1:]:
            mask &= merged_df['phase'].str.contains(pattern, case=False, na=False)
        
        task_frames = merged_df[mask]
        
        if len(task_frames) > 0:
            accuracy = (task_frames['within_radius'].sum() / len(task_frames)) * 100
            mean_distance = task_frames['euclidean_distance'].mean()
            task_breakdown[task_name] = {
                'accuracy': float(accuracy),
                'mean_distance': float(mean_distance),
                'frame_count': len(task_frames)
            }
        else:
            task_breakdown[task_name] = {
                'accuracy': 0.0,
                'mean_distance': np.nan,
                'frame_count': 0
            }
    
    return task_breakdown


def calculate_task_specific_accuracy(merged_df):
    """
    Calculate accuracy breakdown by task type (movement pattern).
    
    Args:
        merged_df: Merged dataframe with 'phase' and 'within_radius' columns
    
    Returns:
        dict with accuracy per task type
    """
    task_breakdown = {}
    
    # Define task type patterns
    task_types = {
        'Horizontal Smooth Pursuit': ['HORIZONTAL'],
        'Vertical Smooth Pursuit': ['VERTICAL'],
        'Circular Smooth Pursuit': ['CIRCULAR'],
        'Structured Saccades': ['SACCADES', 'STRUCTURED'],
        'Random Saccades': ['SACCADES', 'RANDOM']
    }
    
    for task_name, patterns in task_types.items():
        # Filter frames matching all patterns
        mask = merged_df['phase'].str.contains(patterns[0], case=False, na=False)
        for pattern in patterns[1:]:
            mask &= merged_df['phase'].str.contains(pattern, case=False, na=False)
        
        task_frames = merged_df[mask]
        
        if len(task_frames) > 0:
            accuracy = (task_frames['within_radius'].sum() / len(task_frames)) * 100
            mean_distance = task_frames['euclidean_distance'].mean()
            task_breakdown[task_name] = {
                'accuracy': float(accuracy),
                'mean_distance': float(mean_distance),
                'frame_count': len(task_frames)
            }
        else:
            task_breakdown[task_name] = {
                'accuracy': 0.0,
                'mean_distance': np.nan,
                'frame_count': 0
            }
    
    return task_breakdown


def calculate_metrics(detected_csv, gt_csv, offset_frames=0):
    """
    Calculate detection metrics by comparing detected positions with ground truth.
    Only counts errors when GT is outside the detection radius (75px).
    
    Args:
        detected_csv: Path to detected gaze CSV
        gt_csv: Path to ground truth CSV
        offset_frames: Frame offset to apply to detected data
    
    Returns:
        dict with metrics
    """
    # Load data
    detected_df = pd.read_csv(detected_csv)
    gt_df = pd.read_csv(gt_csv)
    
    # Apply offset
    detected_df['frame_adjusted'] = detected_df['frame'] - offset_frames
    
    # Merge with GT
    merged = pd.merge(
        gt_df,
        detected_df,
        left_on='frame',
        right_on='frame_adjusted',
        how='inner'
    )
    
    # Filter for valid GT frames only
    merged_valid = merged[merged['gt_x_px'].notna()].copy()
    
    if len(merged_valid) == 0:
        return {
            'detection_rate': 0.0,
            'mean_distance': np.inf,
            'std_distance': 0.0,
            'max_distance': np.inf,
            'mse': np.inf,
            'rmse': np.inf,
            'fallback_count': 0,
            'frames_within_radius': 0,
            'frames_outside_radius': 0,
            'accuracy': 0.0,
            'deviation_count': 0,
            'threshold_detection_pass': False,
            'threshold_accuracy_pass': False,
            'task_accuracy': {}
        }
    
    # Calculate euclidean distances
    merged_valid['euclidean_distance'] = np.sqrt(
        (merged_valid['x'] - merged_valid['gt_x_px'])**2 +
        (merged_valid['y'] - merged_valid['gt_y_px'])**2
    )
    
    # Classify frames: within radius (correct tracking) vs outside radius (error tracking)
    merged_valid['within_radius'] = merged_valid['euclidean_distance'] <= DISTANCE_THRESHOLD
    merged_valid['is_deviation'] = merged_valid['euclidean_distance'] > DISTANCE_THRESHOLD
    
    # Separate frames for analysis
    frames_within_radius = merged_valid[merged_valid['within_radius']].copy()
    frames_outside_radius = merged_valid[~merged_valid['within_radius']].copy()
    
    # Calculate detection rate (based on all frames)
    detection_rate = (1 - merged_valid['is_interpolated'].sum() / len(merged_valid)) * 100
    fallback_count = int(merged_valid['is_interpolated'].sum())
    
    # Calculate distance metrics ONLY from frames outside radius
    if len(frames_outside_radius) > 0:
        mean_distance = frames_outside_radius['euclidean_distance'].mean()
        std_distance = frames_outside_radius['euclidean_distance'].std()
        max_distance = frames_outside_radius['euclidean_distance'].max()
        # Calculate MSE/RMSE from frames outside radius
        squared_errors = frames_outside_radius['euclidean_distance'] ** 2
        mse = squared_errors.mean()
        rmse = np.sqrt(mse)
    else:
        # All frames within radius = perfect tracking
        mean_distance = 0.0
        std_distance = 0.0
        max_distance = 0.0
        mse = 0.0
        rmse = 0.0
    
    # Accuracy: percentage of frames within radius (correct tracking)
    accuracy = (len(frames_within_radius) / len(merged_valid)) * 100
    deviation_count = len(frames_outside_radius)
    
    # Task-specific accuracy breakdown
    task_accuracy = calculate_task_specific_accuracy(merged_valid)
    
    return {
        'detection_rate': float(detection_rate),
        'mean_distance': float(mean_distance),
        'std_distance': float(std_distance),
        'max_distance': float(max_distance),
        'mse': float(mse),
        'rmse': float(rmse),
        'fallback_count': fallback_count,
        'frames_within_radius': len(frames_within_radius),
        'frames_outside_radius': len(frames_outside_radius),
        'accuracy': float(accuracy),
        'deviation_count': deviation_count,
        'threshold_detection_pass': detection_rate >= DETECTION_THRESHOLD,
        'threshold_accuracy_pass': accuracy >= DETECTION_THRESHOLD,
        'task_accuracy': task_accuracy,  # Task-specific breakdown
        'merged_df': merged_valid  # For plotting (includes all frames with classification)
    }


print("Processing and analysis functions loaded.")


# ==============================================================================
# STATISTICAL TESTING
# ==============================================================================

def perform_statistical_tests(all_metrics):
    """
    Perform ANOVA and pairwise t-tests with Bonferroni correction.
    Uses distances and squared errors from frames outside radius only.
    
    Args:
        all_metrics: dict of {method: {'variant1': metrics, 'variant2': ...}}
    
    Returns:
        DataFrame with statistical test results
    """
    print("\nPerforming statistical tests...")
    
    # Collect euclidean distances and squared errors per method (only frames outside radius)
    distances_by_method = {}
    squared_errors_by_method = {}
    for method in METHODS:
        if method in all_metrics:
            distances = []
            squared_errors = []
            for variant in VARIANTS:
                if variant in all_metrics[method]:
                    # Get merged_df if available
                    if 'merged_df' in all_metrics[method][variant]:
                        df = all_metrics[method][variant]['merged_df']
                        # Only include frames outside radius for error calculation
                        outside_radius = df[~df['within_radius']]
                        distances.extend(outside_radius['euclidean_distance'].tolist())
                        squared_errors.extend((outside_radius['euclidean_distance'] ** 2).tolist())
            distances_by_method[method] = distances
            squared_errors_by_method[method] = squared_errors
    
    # ANOVA on Euclidean Distance
    distance_arrays = [distances_by_method[m] for m in METHODS if m in distances_by_method and len(distances_by_method[m]) > 0]
    
    if len(distance_arrays) < 2:
        print("Warning: Not enough data for statistical tests")
        return pd.DataFrame()
    
    f_stat_dist, p_value_anova_dist = scipy.stats.f_oneway(*distance_arrays)
    
    # ANOVA on MSE (Squared Errors)
    squared_error_arrays = [squared_errors_by_method[m] for m in METHODS if m in squared_errors_by_method and len(squared_errors_by_method[m]) > 0]
    f_stat_mse, p_value_anova_mse = scipy.stats.f_oneway(*squared_error_arrays)
    
    print(f"  ANOVA (Distance): F={f_stat_dist:.4f}, p={p_value_anova_dist:.6f}")
    print(f"  ANOVA (MSE): F={f_stat_mse:.4f}, p={p_value_anova_mse:.6f}")
    
    # Pairwise t-tests with Bonferroni correction
    num_comparisons = len(list(combinations(METHODS, 2)))
    alpha_bonferroni = 0.05 / num_comparisons
    
    pairwise_results = []
    for method_a, method_b in combinations(METHODS, 2):
        if method_a in distances_by_method and method_b in distances_by_method:
            distances_a = distances_by_method[method_a]
            distances_b = distances_by_method[method_b]
            squared_errors_a = squared_errors_by_method[method_a]
            squared_errors_b = squared_errors_by_method[method_b]
            
            if len(distances_a) > 0 and len(distances_b) > 0:
                # Test on Euclidean Distance
                t_stat_dist, p_value_dist = scipy.stats.ttest_ind(distances_a, distances_b)
                p_adjusted_dist = p_value_dist * num_comparisons
                
                # Test on MSE (Squared Errors)
                t_stat_mse, p_value_mse = scipy.stats.ttest_ind(squared_errors_a, squared_errors_b)
                p_adjusted_mse = p_value_mse * num_comparisons
                
                significance_dist = ""
                if p_adjusted_dist < 0.01:
                    significance_dist = "**"
                elif p_adjusted_dist < 0.05:
                    significance_dist = "*"
                
                significance_mse = ""
                if p_adjusted_mse < 0.01:
                    significance_mse = "**"
                elif p_adjusted_mse < 0.05:
                    significance_mse = "*"
                
                pairwise_results.append({
                    'Method_A': method_a,
                    'Method_B': method_b,
                    't_stat_distance': t_stat_dist,
                    'p_value_distance': p_value_dist,
                    'p_adjusted_distance': min(p_adjusted_dist, 1.0),
                    'significant_distance': significance_dist,
                    't_stat_mse': t_stat_mse,
                    'p_value_mse': p_value_mse,
                    'p_adjusted_mse': min(p_adjusted_mse, 1.0),
                    'significant_mse': significance_mse
                })
    
    results_df = pd.DataFrame(pairwise_results)
    results_df['ANOVA_F_distance'] = f_stat_dist
    results_df['ANOVA_p_distance'] = p_value_anova_dist
    results_df['ANOVA_F_mse'] = f_stat_mse
    results_df['ANOVA_p_mse'] = p_value_anova_mse
    results_df['Bonferroni_alpha'] = alpha_bonferroni
    
    print(f"  Bonferroni correction: α = 0.05 / {num_comparisons} = {alpha_bonferroni:.4f}")
    print(f"  Significant comparisons (Distance): {len([r for r in pairwise_results if r['significant_distance']])}")
    print(f"  Significant comparisons (MSE): {len([r for r in pairwise_results if r['significant_mse']])}")
    
    return results_df


print("Statistical testing functions loaded.")


# ==============================================================================
# PLOTTING FUNCTIONS
# ==============================================================================

def get_task_boundaries(gt_csv_path):
    """Extract task boundaries from ground truth CSV."""
    df = pd.read_csv(gt_csv_path)
    phase_changes = df[df['phase'] != df['phase'].shift()].copy()
    
    # Filter to main tasks only (skip instructions/preparations)
    main_tasks = []
    for _, row in phase_changes.iterrows():
        phase = row['phase']
        if any(x in phase for x in ['TUTORIAL', 'TASK_1', 'TASK_2', 'TASK_3', 'TASK_4', 'TASK_5']):
            if 'INSTRUCTION' not in phase and 'PREPARATION' not in phase:
                label = phase.replace('TASK_', 'T').replace('_HORIZONTAL_', ' H-').replace('_VERTICAL_', ' V-')
                label = label.replace('_CIRCULAR_', ' C-').replace('_SACCADES_', ' S-').replace('_LTR', 'L→R')
                label = label.replace('_RTL', 'R→L').replace('_TTB', 'T→B').replace('_BTT', 'B→T')
                label = label.replace('_CW', 'CW').replace('_CCW', 'CCW').replace('_STRUCTURED', 'Struct')
                label = label.replace('_RANDOM', 'Rand').replace('_POINT_', 'P')
                main_tasks.append({'frame': int(row['frame']), 'label': label})
    
    return main_tasks

def plot_error_timeline(metrics_data, method, variant, sync_mode, output_path):
    """
    Plot euclidean distance over time with deviation zones.
    Only frames outside radius are counted in mean calculation.
    
    Args:
        metrics_data: dict containing 'merged_df' with euclidean_distance column
        method: Method name
        variant: Variant name
        sync_mode: 'no_sync' or 'auto_sync'
        output_path: Path to save plot
    """
    if 'merged_df' not in metrics_data or len(metrics_data['merged_df']) == 0:
        print(f"    Warning: No data to plot for {method}_{variant}")
        return
    
    df = metrics_data['merged_df']
    
    # Determine which frame column to use (after merge, GT frame column is the one we want)
    # If there's a 'frame' column directly, use it; otherwise use index
    if 'frame' in df.columns:
        frame_col = df['frame']
    elif 'frame_x' in df.columns:  # pandas adds _x suffix to left frame
        frame_col = df['frame_x']
    else:
        frame_col = df.index  # fallback to index
    
    # Get task boundaries
    task_boundaries = get_task_boundaries(GT_CSV_PATH)
    
    fig, ax = plt.subplots(figsize=(14, 6))
    
    # Plot distance with color zones: green = within radius, red = outside radius
    colors = ['green' if within else 'red' for within in df['within_radius']]
    ax.scatter(frame_col, df['euclidean_distance'], c=colors, s=1, alpha=0.5)
    
    ax.axhline(y=DISTANCE_THRESHOLD, color='blue', linestyle='--', linewidth=2, label=f'Radius Threshold ({DISTANCE_THRESHOLD} px)')
    
    # Add task boundary lines
    for task in task_boundaries:
        ax.axvline(x=task['frame'], color='gray', linestyle=':', linewidth=1, alpha=0.5)
        ax.text(task['frame'], ax.get_ylim()[1]*0.95, task['label'], 
                rotation=90, va='top', ha='right', fontsize=7, alpha=0.7)
    
    # Annotations - only show mean from frames outside radius
    mean_dist = metrics_data['mean_distance']
    std_dist = metrics_data['std_distance']
    accuracy = metrics_data['accuracy']
    frames_outside = metrics_data['frames_outside_radius']
    frames_within = metrics_data['frames_within_radius']
    
    info_text = f"Mean Error (outside radius): {mean_dist:.2f} ± {std_dist:.2f} px\n"
    info_text += f"Accuracy (within radius): {accuracy:.1f}%\n"
    info_text += f"Frames: {frames_within} within / {frames_outside} outside"
    
    ax.text(0.02, 0.98, info_text,
            transform=ax.transAxes, fontsize=10, verticalalignment='top',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
    
    ax.set_xlabel('Frame Number', fontsize=12)
    ax.set_ylabel('Euclidean Distance (pixels)', fontsize=12)
    ax.set_title(f'{method.upper()} - {variant} - Distance Timeline ({sync_mode})', fontsize=14)
    ax.legend()
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    plt.savefig(output_path, dpi=150)
    plt.close(fig)


def plot_trajectory_overlay(metrics_data, method, variant, sync_mode, output_path):
    """
    Plot detected vs ground truth trajectories.
    
    Args:
        metrics_data: dict containing 'merged_df' with coordinates
        method: Method name
        variant: Variant name
        sync_mode: 'no_sync' or 'auto_sync'
        output_path: Path to save plot
    """
    if 'merged_df' not in metrics_data or len(metrics_data['merged_df']) == 0:
        print(f"    Warning: No data to plot for {method}_{variant}")
        return
    
    df = metrics_data['merged_df']
    
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Plot GT (red) and detected (blue)
    ax.scatter(df['gt_x_px'], df['gt_y_px'], c='red', alpha=0.3, s=10, label='Ground Truth')
    ax.scatter(df['x'], df['y'], c='blue', alpha=0.3, s=10, label='Detected')
    
    ax.set_xlabel('X Coordinate (pixels)', fontsize=12)
    ax.set_ylabel('Y Coordinate (pixels)', fontsize=12)
    ax.set_title(f'{method.upper()} - {variant} - Trajectory Overlay ({sync_mode})', fontsize=14)
    ax.legend()
    ax.grid(True, alpha=0.3)
    ax.set_xlim(0, WIDTH)
    ax.set_ylim(0, HEIGHT)
    ax.invert_yaxis()  # Match image coordinates
    
    plt.tight_layout()
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    plt.savefig(output_path, dpi=150)
    plt.close(fig)


print("Plotting functions loaded.")


# ==============================================================================
# VIDEO OVERLAY GENERATION
# ==============================================================================

def generate_overlay_video(video_path, detected_csv, gt_csv, method, variant, sync_mode, offset_frames, output_path):
    """
    Generate overlay video showing detected vs ground truth gaze positions.
    
    Args:
        video_path: Path to input video
        detected_csv: Path to detected gaze CSV
        gt_csv: Path to ground truth CSV
        method: Method name
        variant: Variant name
        sync_mode: 'no_sync' or 'auto_sync'
        offset_frames: Frame offset to apply
        output_path: Path to save overlay video
    """
    print(f"    Generating overlay video: {method}_{variant}_{sync_mode}...")
    
    # Load detection and GT data
    detected_df = pd.read_csv(detected_csv)
    gt_df = pd.read_csv(gt_csv)
    
    # Apply offset
    detected_df['frame_adjusted'] = detected_df['frame'] - offset_frames
    
    # Merge
    merged = pd.merge(
        gt_df,
        detected_df,
        left_on='frame',
        right_on='frame_adjusted',
        how='inner'
    )
    merged_valid = merged[merged['gt_x_px'].notna()].copy()
    
    if len(merged_valid) == 0:
        print(f"      Warning: No valid frames to render for {method}_{variant}")
        return
    
    # Calculate distances for color coding
    merged_valid['euclidean_distance'] = np.sqrt(
        (merged_valid['x'] - merged_valid['gt_x_px'])**2 +
        (merged_valid['y'] - merged_valid['gt_y_px'])**2
    )
    merged_valid['is_accurate'] = merged_valid['euclidean_distance'] <= DISTANCE_THRESHOLD
    merged_valid['within_radius'] = merged_valid['euclidean_distance'] <= DISTANCE_THRESHOLD
    
    # Open video
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print(f"      Error: Cannot open video {video_path}")
        return
    
    # Get video properties
    width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS)
    if fps == 0 or fps is None:
        fps = 60  # Default to 60 FPS if unable to read
    
    # Create video writer
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(output_path, fourcc, fps, (width, height))
    
    # Verify video writer opened successfully
    if not out.isOpened():
        print(f"      Error: Cannot create video writer for {output_path}")
        cap.release()
        return
    
    # Get frame range
    min_frame = int(merged_valid['frame'].min())
    max_frame = int(merged_valid['frame'].max())
    
    # Create lookup dict for fast access
    frame_data = {}
    for _, row in merged_valid.iterrows():
        frame_data[int(row['frame'])] = row
    
    # Process ALL frames from start to render complete video
    current_frame = 0
    frames_written = 0
    pbar = tqdm(total=max_frame + 1, desc=f"      Rendering", unit="frame", leave=False)
    
    while current_frame <= max_frame:
        ret, frame = cap.read()
        if not ret:
            break
        
        # Only draw overlays for frames we have data for
        if current_frame in frame_data and current_frame >= min_frame:
            row = frame_data[current_frame]
            
            # Validate coordinates are within bounds
            gt_x, gt_y = int(row['gt_x_px']), int(row['gt_y_px'])
            det_x, det_y = int(row['x']), int(row['y'])
            
            # Clamp coordinates to frame boundaries
            gt_x = max(0, min(width - 1, gt_x))
            gt_y = max(0, min(height - 1, gt_y))
            det_x = max(0, min(width - 1, det_x))
            det_y = max(0, min(height - 1, det_y))
            
            # Calculate distance for visualization
            distance = row['euclidean_distance']
            is_accurate = row['is_accurate']
            
            # Create overlay for semi-transparent elements
            overlay = frame.copy()
            
            # 1. Draw threshold radius circle around GT (shows acceptable zone)
            threshold_color = (0, 255, 0) if is_accurate else (0, 165, 255)  # Green if within, Orange if outside
            cv2.circle(overlay, (gt_x, gt_y), int(DISTANCE_THRESHOLD), threshold_color, 2)
            
            # 2. Draw distance radius circle around GT (shows actual error magnitude)
            if distance > 5:  # Only draw if distance is significant
                distance_color = (0, 255, 0) if is_accurate else (0, 0, 255)
                cv2.circle(overlay, (gt_x, gt_y), int(distance), distance_color, 1)
            
            # Apply transparency to circles
            cv2.addWeighted(overlay, 0.3, frame, 0.7, 0, frame)
            
            # 3. Draw ground truth (green circle with label)
            cv2.circle(frame, (gt_x, gt_y), 12, (0, 255, 0), 2)
            cv2.circle(frame, (gt_x, gt_y), 3, (0, 255, 0), -1)
            
            # 4. Draw detected (color based on accuracy)
            det_color = (0, 255, 255) if is_accurate else (0, 0, 255)  # Yellow if accurate, Red if not
            cv2.circle(frame, (det_x, det_y), 12, det_color, 2)
            cv2.circle(frame, (det_x, det_y), 3, det_color, -1)
            
            # 5. Draw connecting line with distance annotation
            cv2.line(frame, (gt_x, gt_y), (det_x, det_y), (255, 255, 255), 2)
            
            # Draw distance text at midpoint of line
            mid_x = (gt_x + det_x) // 2
            mid_y = (gt_y + det_y) // 2
            dist_text = f"{distance:.1f}px"
            # Add background rectangle for better readability
            (text_width, text_height), baseline = cv2.getTextSize(dist_text, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)
            cv2.rectangle(frame, (mid_x - 5, mid_y - text_height - 5), 
                         (mid_x + text_width + 5, mid_y + 5), (0, 0, 0), -1)
            cv2.putText(frame, dist_text, (mid_x, mid_y), cv2.FONT_HERSHEY_SIMPLEX, 
                       0.5, (255, 255, 255), 1, cv2.LINE_AA)
            
            # 6. Draw main distance info in top-left
            text_color = (0, 255, 255) if is_accurate else (0, 0, 255)
            status = "WITHIN THRESHOLD" if is_accurate else "EXCEEDS THRESHOLD"
            cv2.putText(frame, f"Distance: {distance:.1f}px - {status}", 
                       (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 0.7, text_color, 2)
            
            # 7. Draw distance scale bar (horizontal bar showing distance relative to threshold)
            bar_x, bar_y = 10, 70
            bar_width = 300
            bar_height = 20
            # Background bar
            cv2.rectangle(frame, (bar_x, bar_y), (bar_x + bar_width, bar_y + bar_height), (50, 50, 50), -1)
            # Threshold marker
            threshold_pos = bar_x + bar_width
            cv2.line(frame, (threshold_pos, bar_y - 5), (threshold_pos, bar_y + bar_height + 5), (0, 255, 0), 2)
            cv2.putText(frame, f"{int(DISTANCE_THRESHOLD)}px", (threshold_pos - 20, bar_y - 10), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 0), 1)
            # Current distance bar
            distance_ratio = min(distance / DISTANCE_THRESHOLD, 1.5)  # Cap at 150% for display
            distance_bar_width = int(bar_width * distance_ratio)
            bar_color = (0, 255, 0) if is_accurate else (0, 0, 255)
            cv2.rectangle(frame, (bar_x, bar_y), (bar_x + distance_bar_width, bar_y + bar_height), bar_color, -1)
            
            # 8. Draw legend
            legend_y = bar_y + bar_height + 25
            cv2.putText(frame, "GT (Ground Truth)", (10, legend_y), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
            cv2.putText(frame, "Detected", (10, legend_y + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, det_color, 1)
            cv2.putText(frame, f"Threshold: {int(DISTANCE_THRESHOLD)}px", (10, legend_y + 40), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (128, 128, 128), 1)
            cv2.putText(frame, f"Frame: {current_frame}", (10, height - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255, 255, 255), 1)
        
        # Write frame regardless of whether we have data (to maintain video continuity)
        if current_frame >= min_frame:
            out.write(frame)
            frames_written += 1
        
        pbar.update(1)
        current_frame += 1
    
    pbar.close()
    cap.release()
    out.release()
    
    # Verify the output file was actually created and has content
    if os.path.exists(output_path):
        file_size = os.path.getsize(output_path)
        if file_size > 0:
            print(f"      ✓ Overlay video saved: {os.path.basename(output_path)} ({frames_written} frames, {file_size / (1024*1024):.1f} MB)")
        else:
            print(f"      ⚠ Warning: Video file created but is empty: {os.path.basename(output_path)}")
    else:
        print(f"      ✗ Error: Failed to create video file: {os.path.basename(output_path)}")


print("Video overlay functions loaded.")


# ==============================================================================
# EXCEL REPORT GENERATION
# ==============================================================================

def generate_excel_report(all_metrics, all_offsets, statistical_results, sync_mode, output_path):
    """
    Generate comprehensive Excel report with multiple sheets.
    
    Args:
        all_metrics: Nested dict {method: {variant: metrics_dict}}
        all_offsets: Dict {method_{variant}: offset_dict} (for auto_sync only)
        statistical_results: DataFrame from perform_statistical_tests
        sync_mode: 'no_sync' or 'auto_sync'
        output_path: Path to save Excel file
    """
    print(f"\nGenerating Excel report for {sync_mode}...")
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # ===== SHEET 1: SUMMARY =====
        summary_data = []
        for method in METHODS:
            if method in all_metrics:
                detection_rates = [all_metrics[method][v]['detection_rate'] for v in VARIANTS if v in all_metrics[method]]
                accuracies = [all_metrics[method][v]['accuracy'] for v in VARIANTS if v in all_metrics[method]]
                distances = [all_metrics[method][v]['mean_distance'] for v in VARIANTS if v in all_metrics[method]]
                mses = [all_metrics[method][v]['mse'] for v in VARIANTS if v in all_metrics[method]]
                rmses = [all_metrics[method][v]['rmse'] for v in VARIANTS if v in all_metrics[method]]
                
                if detection_rates and accuracies and distances:
                    summary_data.append({
                        'Method': method,
                        'Detection_Rate_Mean': np.mean(detection_rates),
                        'Accuracy_Mean': np.mean(accuracies),
                        'Distance_Mean': np.mean(distances),
                        'MSE_Mean': np.mean(mses),
                        'RMSE_Mean': np.mean(rmses),
                        'Both_Thresholds_Pass': np.mean(detection_rates) >= DETECTION_THRESHOLD and np.mean(accuracies) >= DETECTION_THRESHOLD
                    })
        
        summary_df = pd.DataFrame(summary_data)
        if len(summary_df) > 0:
            summary_df = summary_df.sort_values('Distance_Mean')
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Add explanation rows
        ws = writer.sheets['Summary']
        last_row = len(summary_df) + 3
        ws.cell(row=last_row, column=1, value="IMPORTANT NOTE:")
        ws.cell(row=last_row+1, column=1, 
                value="Distance_Mean: Average error calculated ONLY from frames where GT is OUTSIDE detection radius (>75px)")
        ws.cell(row=last_row+2, column=1, 
                value="Frames within radius (≤75px) are considered correct tracking and excluded from error calculation")
        ws.cell(row=last_row+3, column=1, value="Statistical Note:")
        ws.cell(row=last_row+4, column=1, 
                value=f"Bonferroni correction applied: α = 0.05 / {len(list(combinations(METHODS, 2)))} = {0.05/len(list(combinations(METHODS, 2))):.4f}")
        
        # ===== SHEET 2-7: DETECTION RATE, ACCURACY, DISTANCE, STD DISTANCE, MSE, RMSE TABLES =====
        for metric_name, metric_key in [('Detection_Rate', 'detection_rate'),
                                         ('Accuracy', 'accuracy'),
                                         ('Mean_Distance', 'mean_distance'),
                                         ('Std_Distance', 'std_distance'),
                                         ('MSE', 'mse'),
                                         ('RMSE', 'rmse')]:
            table_data = []
            for method in METHODS:
                if method in all_metrics:
                    row = {'Method': method}
                    values = []
                    for variant in VARIANTS:
                        if variant in all_metrics[method]:
                            val = all_metrics[method][variant][metric_key]
                            row[variant] = val
                            values.append(val)
                        else:
                            row[variant] = np.nan
                    
                    if values:
                        row['Mean'] = np.mean(values)
                        row['Std'] = np.std(values)
                    table_data.append(row)
            
            table_df = pd.DataFrame(table_data)
            table_df.to_excel(writer, sheet_name=f'{metric_name}_Table', index=False)
        
        # ===== SHEET 6: ROBUSTNESS RANKING =====
        robustness_data = []
        for method in METHODS:
            if method in all_metrics:
                detection_rates = [all_metrics[method][v]['detection_rate'] for v in VARIANTS if v in all_metrics[method]]
                accuracies = [all_metrics[method][v]['accuracy'] for v in VARIANTS if v in all_metrics[method]]
                distances = [all_metrics[method][v]['mean_distance'] for v in VARIANTS if v in all_metrics[method]]
                mses = [all_metrics[method][v]['mse'] for v in VARIANTS if v in all_metrics[method]]
                rmses = [all_metrics[method][v]['rmse'] for v in VARIANTS if v in all_metrics[method]]
                
                if detection_rates and accuracies and distances:
                    dr_mean, dr_std = np.mean(detection_rates), np.std(detection_rates)
                    acc_mean, acc_std = np.mean(accuracies), np.std(accuracies)
                    dist_mean, dist_std = np.mean(distances), np.std(distances)
                    mse_mean, mse_std = np.mean(mses), np.std(mses)
                    rmse_mean, rmse_std = np.mean(rmses), np.std(rmses)
                    
                    robustness_data.append({
                        'Method': method,
                        'Detection_Rate_CV': dr_std / dr_mean if dr_mean > 0 else np.inf,
                        'Accuracy_CV': acc_std / acc_mean if acc_mean > 0 else np.inf,
                        'Distance_CV': dist_std / dist_mean if dist_mean > 0 else np.inf,
                        'MSE_CV': mse_std / mse_mean if mse_mean > 0 else np.inf,
                        'RMSE_CV': rmse_std / rmse_mean if rmse_mean > 0 else np.inf
                    })
        
        robustness_df = pd.DataFrame(robustness_data)
        if len(robustness_df) > 0:
            robustness_df = robustness_df.sort_values('Distance_CV')
        robustness_df.to_excel(writer, sheet_name='Robustness_Ranking', index=False)
        
        # ===== SHEET 7: NATURAL DELAY (auto_sync only) =====
        if sync_mode == 'auto_sync' and all_offsets:
            delay_data = []
            for key, offset_info in all_offsets.items():
                method, variant = key.split('_', 1)
                delay_data.append({
                    'Method': method,
                    'Variant': variant,
                    'Offset_Frames': offset_info['offset_frames'],
                    'Delay_ms': offset_info['delay_ms'],
                    'Correlation_Coef': offset_info['correlation_coef']
                })
            
            delay_df = pd.DataFrame(delay_data)
            delay_df.to_excel(writer, sheet_name='Natural_Delay', index=False)
        
        # ===== SHEET 8: THRESHOLD VALIDATION =====
        validation_data = []
        for method in METHODS:
            if method in all_metrics:
                for variant in VARIANTS:
                    if variant in all_metrics[method]:
                        m = all_metrics[method][variant]
                        validation_data.append({
                            'Method': method,
                            'Variant': variant,
                            'Detection_Rate': m['detection_rate'],
                            'Detection_Pass': m['threshold_detection_pass'],
                            'Accuracy': m['accuracy'],
                            'Accuracy_Pass': m['threshold_accuracy_pass'],
                            'Mean_Distance': m['mean_distance'],
                            'Both_Pass': m['threshold_detection_pass'] and m['threshold_accuracy_pass']
                        })
        
        validation_df = pd.DataFrame(validation_data)
        validation_df.to_excel(writer, sheet_name='Threshold_Validation', index=False)
        
        # Apply conditional formatting
        ws = writer.sheets['Threshold_Validation']
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        for row_idx, row in enumerate(validation_df.itertuples(), start=2):
            # Detection Pass
            cell = ws.cell(row=row_idx, column=4)
            cell.fill = green_fill if row.Detection_Pass else red_fill
            
            # Accuracy Pass
            cell = ws.cell(row=row_idx, column=6)
            cell.fill = green_fill if row.Accuracy_Pass else red_fill
            
            # Both Pass
            cell = ws.cell(row=row_idx, column=8)
            cell.fill = green_fill if row.Both_Pass else red_fill
        
        # ===== SHEET 9: STATISTICAL TESTS =====
        if not statistical_results.empty:
            statistical_results.to_excel(writer, sheet_name='Statistical_Tests', index=False)
        
        # ===== SHEETS 10-14: TASK-SPECIFIC ACCURACY =====
        task_names = ['Horizontal Smooth Pursuit', 'Vertical Smooth Pursuit', 
                      'Circular Smooth Pursuit', 'Structured Saccades', 'Random Saccades']
        
        for task_name in task_names:
            task_data = []
            for method in METHODS:
                if method in all_metrics:
                    row = {'Method': method}
                    accuracies = []
                    for variant in VARIANTS:
                        if variant in all_metrics[method]:
                            task_acc = all_metrics[method][variant].get('task_accuracy', {})
                            if task_name in task_acc:
                                acc_val = task_acc[task_name]['accuracy']
                                row[variant] = acc_val
                                accuracies.append(acc_val)
                            else:
                                row[variant] = np.nan
                        else:
                            row[variant] = np.nan
                    
                    if accuracies:
                        row['Mean'] = np.mean(accuracies)
                        row['Std'] = np.std(accuracies)
                    task_data.append(row)
            
            task_df = pd.DataFrame(task_data)
            sheet_name = task_name.replace(' ', '_')[:31]  # Excel sheet name limit
            task_df.to_excel(writer, sheet_name=f'{sheet_name}_Acc', index=False)
        
        # ===== SHEET 15: TASK ACCURACY SUMMARY =====
        summary_task_data = []
        for method in METHODS:
            if method in all_metrics:
                row = {'Method': method}
                for task_name in task_names:
                    task_accuracies = []
                    for variant in VARIANTS:
                        if variant in all_metrics[method]:
                            task_acc = all_metrics[method][variant].get('task_accuracy', {})
                            if task_name in task_acc and task_acc[task_name]['frame_count'] > 0:
                                task_accuracies.append(task_acc[task_name]['accuracy'])
                    
                    col_name = task_name.replace(' Smooth Pursuit', '_SP').replace(' Saccades', '_Sacc')
                    if task_accuracies:
                        row[col_name] = np.mean(task_accuracies)
                    else:
                        row[col_name] = np.nan
                summary_task_data.append(row)
        
        summary_task_df = pd.DataFrame(summary_task_data)
        summary_task_df.to_excel(writer, sheet_name='Task_Accuracy_Summary', index=False)
    
    print(f"  ✓ Excel report saved to: {output_path}")


print("Excel generation functions loaded.")


# ==============================================================================
# COMPARISON ANALYSIS ORCHESTRATOR
# ==============================================================================

def run_comparison_analysis(sync_mode):
    """
    Run complete comparison analysis for a sync mode.
    
    Args:
        sync_mode: 'no_sync' or 'auto_sync'
    """
    print("\n" + "=" * 80)
    print(f"COMPARISON ANALYSIS - {sync_mode.upper()}")
    print("=" * 80)
    
    all_metrics = {}
    all_offsets = {}
    
    # Progress bar for combinations
    total_combos = len(METHODS) * len(VARIANTS)
    pbar = tqdm(total=total_combos, desc="Calculating metrics", unit="combo")
    
    for method in METHODS:
        all_metrics[method] = {}
        
        for variant in VARIANTS:
            combo_name = f"{method}_{variant}"
            detected_csv = os.path.join(DETECTION_RESULTS_DIR, f"{combo_name}.csv")
            
            if not os.path.exists(detected_csv):
                print(f"  Warning: Missing {detected_csv}")
                pbar.update(1)
                continue
            
            # Calculate offset if auto_sync
            offset_frames = 0
            if sync_mode == 'auto_sync':
                offset_info = calculate_offset_limited(detected_csv, GT_CSV_PATH)
                offset_frames = offset_info['offset_frames']
                all_offsets[combo_name] = offset_info
            
            # Calculate metrics
            metrics = calculate_metrics(detected_csv, GT_CSV_PATH, offset_frames)
            all_metrics[method][variant] = metrics
            
            pbar.update(1)
    
    pbar.close()
    
    # Statistical tests
    statistical_results = perform_statistical_tests(all_metrics)
    
    # Generate plots
    print(f"\nGenerating plots for {sync_mode}...")
    plots_dir = os.path.join(PLOTS_DIR, sync_mode)
    plot_count = 0
    
    for method in METHODS:
        if method in all_metrics:
            for variant in VARIANTS:
                if variant in all_metrics[method]:
                    metrics_data = all_metrics[method][variant]
                    
                    # Error timeline
                    timeline_path = os.path.join(plots_dir, f"{method}_{variant}_error_timeline.png")
                    plot_error_timeline(metrics_data, method, variant, sync_mode, timeline_path)
                    
                    # Trajectory overlay
                    trajectory_path = os.path.join(plots_dir, f"{method}_{variant}_trajectory_overlay.png")
                    plot_trajectory_overlay(metrics_data, method, variant, sync_mode, trajectory_path)
                    
                    plot_count += 2
    
    print(f"  ✓ Generated {plot_count} plots")
    
    # Generate overlay videos
    print(f"\nGenerating overlay videos for {sync_mode}...")
    video_dir = os.path.join(VIDEO_OVERLAYS_DIR, sync_mode)
    video_count = 0
    
    for method in METHODS:
        if method in all_metrics:
            for variant in VARIANTS:
                if variant in all_metrics[method]:
                    combo_name = f"{method}_{variant}"
                    detected_csv = os.path.join(DETECTION_RESULTS_DIR, f"{combo_name}.csv")
                    video_path = os.path.join(VIDEO_DIR, f"{variant}.mp4")
                    
                    if os.path.exists(detected_csv) and os.path.exists(video_path):
                        # Get offset if auto_sync
                        offset_frames = 0
                        if sync_mode == 'auto_sync' and combo_name in all_offsets:
                            offset_frames = all_offsets[combo_name]['offset_frames']
                        
                        output_video_path = os.path.join(video_dir, f"{combo_name}_overlay.mp4")
                        
                        try:
                            generate_overlay_video(
                                video_path, detected_csv, GT_CSV_PATH,
                                method, variant, sync_mode, offset_frames,
                                output_video_path
                            )
                            video_count += 1
                        except Exception as e:
                            print(f"      Warning: Failed to generate overlay video for {combo_name}: {e}")
    
    print(f"  ✓ Generated {video_count} overlay videos")
    
    # Generate Excel report
    report_path = os.path.join(REPORTS_DIR, f"comparison_report_{sync_mode}.xlsx")
    generate_excel_report(all_metrics, all_offsets, statistical_results, sync_mode, report_path)
    
    print(f"\n{sync_mode.upper()} analysis complete!")
    return all_metrics, all_offsets, statistical_results


print("Analysis orchestrator loaded.")


# ==============================================================================
# MAIN FUNCTION
# ==============================================================================

def main():
    """Main entry point for the comparison pipeline."""
    print("\n" + "=" * 80)
    print("STARTING COMPARISON PIPELINE")
    print("=" * 80)
    
    # Create output directories
    for directory in [OUTPUT_BASE_DIR, DETECTION_RESULTS_DIR, PLOTS_DIR, 
                      REPORTS_DIR, VIDEO_OVERLAYS_DIR, CHECKPOINTS_DIR, LOGS_DIR]:
        os.makedirs(directory, exist_ok=True)
    
    # Initialize managers
    checkpoint_manager = CheckpointManager()
    error_logger = ErrorLogger()
    
    # Check existing progress
    completed, total = checkpoint_manager.get_progress()
    
    if completed > 0:
        print(f"\nFound existing progress: {completed}/{total} combinations completed")
        print("\nOptions:")
        print("  1. Resume from checkpoint")
        print("  2. Restart (clear checkpoint)")
        print("  3. Skip to analysis (use existing detection results)")
        
        choice = input("\nEnter your choice (1/2/3): ").strip()
        
        if choice == '2':
            print("Clearing checkpoint...")
            checkpoint_manager.clear()
            # Optionally clear detection results
            clear_results = input("Also clear existing detection results? (y/n): ").strip().lower()
            if clear_results == 'y':
                import shutil
                if os.path.exists(DETECTION_RESULTS_DIR):
                    shutil.rmtree(DETECTION_RESULTS_DIR)
                os.makedirs(DETECTION_RESULTS_DIR)
                print("Detection results cleared.")
        elif choice == '3':
            print("Skipping to analysis phase...")
            run_comparison_analysis('no_sync')
            run_comparison_analysis('auto_sync')
            print("\n" + "=" * 80)
            print("PIPELINE COMPLETED SUCCESSFULLY!")
            print("=" * 80)
            return
    
    # Validate videos
    video_properties = validate_video_properties()
    
    if len(video_properties) != len(VARIANTS):
        print("\nError: Not all video variants are available!")
        return
    
    # Run detection phase
    run_detection_phase(checkpoint_manager, error_logger)
    
    # Run comparison analysis (both sync modes)
    run_comparison_analysis('no_sync')
    run_comparison_analysis('auto_sync')
    
    print("\n" + "=" * 80)
    print("PIPELINE COMPLETED SUCCESSFULLY!")
    print("=" * 80)
    print(f"\nResults saved to: {OUTPUT_BASE_DIR}")
    print(f"  - Detection CSVs: {DETECTION_RESULTS_DIR}")
    print(f"  - Plots: {PLOTS_DIR}")
    print(f"  - Overlay Videos: {VIDEO_OVERLAYS_DIR}")
    print(f"  - Excel Reports: {REPORTS_DIR}")
    print(f"  - Logs: {LOGS_DIR}")


if __name__ == "__main__":
    main()
