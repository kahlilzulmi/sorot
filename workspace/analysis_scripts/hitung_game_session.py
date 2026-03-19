# Import libraries
import pandas as pd
import os
import glob
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from pathlib import Path
from tqdm import tqdm
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ==============================================================================
# CONFIGURATION
# ==============================================================================
# Path to the Game Session directory
# Using raw string for Windows path with spaces
default_data_dir = r"Game Session"
DATA_DIR = input(f"Masukkan path Game Session (default: {default_data_dir}): ").strip()
if not DATA_DIR:
    DATA_DIR = default_data_dir

if not os.path.exists(DATA_DIR):
    print(f"Error: Directory not found: {DATA_DIR}")
    exit()

# Output directory for charts and Excel
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_OUTPUT_DIR = os.path.join(OUTPUT_DIR, "game_output")

# Create output directory if it doesn't exist
os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)

# ==============================================================================
# DATA LOADING & PROCESSING
# ==============================================================================
print(f"Scanning for game sessions in: {DATA_DIR}")

session_durations = []
all_gaze_x = []
all_gaze_y = []
all_gaze_u = []  # normalized X (0..1)
all_gaze_v = []  # normalized Y (0..1)
screen_resolution = (1920, 1080) # Default fallback
incomplete_sessions = []
gaze_point_stats = []
session_metadata_list = []

# Find all subdirectories (assuming each session is a folder)
# If the structure is flat CSVs, we'll adjust. 
# Based on context, it seems to be folders like 'game_session_YYYY-MM-DD...'
session_folders = glob.glob(os.path.join(DATA_DIR, "game_session_*"))

print(f"Found {len(session_folders)} potential session folders.")

valid_sessions_count = 0

for session_path in tqdm(session_folders, desc="Processing Sessions"):
    session_name = os.path.basename(session_path)
    
    # Define file paths
    results_path = os.path.join(session_path, "quiz_results.csv")
    gaze_path = os.path.join(session_path, "gaze_data.csv")
    metadata_path = os.path.join(session_path, "session_metadata.csv")
    
    # 1. Check if session is complete (6 questions answered)
    if not os.path.exists(results_path):
        incomplete_sessions.append({'session_id': session_name, 'reason': 'Missing quiz_results.csv'})
        continue
        
    try:
        results_df = pd.read_csv(results_path)
        # Assuming one row per question. Check if we have 6 rows.
        if len(results_df) < 6:
            incomplete_sessions.append({'session_id': session_name, 'reason': f'Incomplete - only {len(results_df)} questions'})
            continue
    except Exception as e:
        incomplete_sessions.append({'session_id': session_name, 'reason': f'Error reading results: {e}'})
        print(f"Error reading results for {session_name}: {e}")
        continue

    # 2. Load Metadata for screen resolution (update if available)
    metadata_dict = {'session_id': session_name}
    # Default per-session resolution; do not rely on previous session's value
    session_resolution = (1920, 1080)
    if os.path.exists(metadata_path):
        try:
            meta_df = pd.read_csv(metadata_path)
            if not meta_df.empty:
                # Store all metadata except screen_resolution
                for col in meta_df.columns:
                    if col.lower() != 'screen_resolution':
                        metadata_dict[col] = meta_df.iloc[0][col]
                
                # Still use screen_resolution for processing but don't store it
                if 'screen_resolution' in meta_df.columns:
                    res_str = meta_df.iloc[0]['screen_resolution']
                    # Parse "1920x1080"
                    w, h = map(int, res_str.lower().split('x'))
                    screen_resolution = (w, h)  # keep last seen as fallback for pixel-based chart
                    session_resolution = (w, h)  # use this resolution for normalization in this session
        except Exception as e:
            pass # Keep default or previous
    
    session_metadata_list.append(metadata_dict)

    # 3. Process Gaze Data
    if os.path.exists(gaze_path):
        try:
            gaze_df = pd.read_csv(gaze_path)
            
            # Calculate gaze point statistics
            total_rows = len(gaze_df)
            valid_gaze = gaze_df.dropna(subset=['gaze_x', 'gaze_y'])
            valid_count = len(valid_gaze)
            data_quality = (valid_count / total_rows * 100) if total_rows > 0 else 0
            
            # Collect gaze points for Heatmap
            all_gaze_x.extend(valid_gaze['gaze_x'].tolist())
            all_gaze_y.extend(valid_gaze['gaze_y'].tolist())

            # Collect normalized gaze points (relative to screen size 0..1)
            if session_resolution[0] > 0 and session_resolution[1] > 0:
                norm_x = (valid_gaze['gaze_x'] / session_resolution[0]).clip(0, 1)
                norm_y = (valid_gaze['gaze_y'] / session_resolution[1]).clip(0, 1)
                all_gaze_u.extend(norm_x.tolist())
                all_gaze_v.extend(norm_y.tolist())
            
            # Calculate average gaze points per question
            question_gaze_counts = gaze_df[gaze_df['question_index'] >= 0].groupby('question_index').size()
            avg_gaze_per_question = question_gaze_counts.mean() if not question_gaze_counts.empty else 0
            
            gaze_point_stats.append({
                'session_id': session_name,
                'total_gaze_points': total_rows,
                'valid_gaze_points': valid_count,
                'avg_gaze_points_per_question': avg_gaze_per_question,
                'data_quality_score': round(data_quality, 2)
            })
            
            # Calculate Duration per Question
            # Group by 'question_index'
            if 'question_index' in gaze_df.columns and 'timestamp' in gaze_df.columns:
                # Filter out non-question phases if necessary (e.g. question_index < 0)
                # Assuming question_index 0-5 are the questions
                # Ensure numeric dtype for question_index
                gaze_df['question_index'] = pd.to_numeric(gaze_df['question_index'], errors='coerce')
                question_groups = gaze_df[gaze_df['question_index'] >= 0].groupby('question_index')
                
                for q_idx, group in question_groups:
                    if group.empty:
                        continue
                    
                    # Duration = max timestamp - min timestamp
                    duration = group['timestamp'].max() - group['timestamp'].min()
                    
                    # Robustly convert question index to int using group's first value
                    q_val = pd.to_numeric(group['question_index'].iloc[0], errors='coerce')
                    if pd.isna(q_val):
                        continue
                    q_num = int(q_val)

                    session_durations.append({
                        'session_id': session_name,
                        'question_id': q_num + 1, # 1-based index for display
                        'gaze_duration': duration
                    })
            
            valid_sessions_count += 1
            
        except Exception as e:
            print(f"Error reading gaze data for {session_name}: {e}")
            incomplete_sessions.append({'session_id': session_name, 'reason': f'Error reading gaze data: {e}'})

print(f"Successfully processed {valid_sessions_count} valid sessions.")

if not session_durations:
    print("No valid data found. Exiting.")
    exit()

# Create DataFrame for analysis
df_durations = pd.DataFrame(session_durations)

# Sort by session_id and question_id
df_durations = df_durations.sort_values(['session_id', 'question_id']).reset_index(drop=True)

# ==============================================================================
# EXCEL EXPORT
# ==============================================================================
print("\n--- Exporting to Excel ---")

# Generate timestamp for filename
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
excel_filename = f"game_session_analysis_{timestamp}.xlsx"
excel_path = os.path.join(EXCEL_OUTPUT_DIR, excel_filename)

# Create Excel writer
with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    
    # Sheet 1: Session Details (raw data)
    df_durations.to_excel(writer, sheet_name='Session Details', index=False)
    
    # Sheet 2: Per-Question Summary
    per_question_stats = (
        df_durations.groupby('question_id')['gaze_duration']
        .agg(['count', 'mean', 'std', 'min', 'max'])
        .reset_index()
    )
    per_question_stats.to_excel(writer, sheet_name='Per-Question Summary', index=False)
    
    # Sheet 3: Overall Statistics
    overall_stats = pd.DataFrame({
        'Statistic': ['Min', 'Max', 'Mean', 'Std'],
        'Value': [
            df_durations['gaze_duration'].min(),
            df_durations['gaze_duration'].max(),
            df_durations['gaze_duration'].mean(),
            df_durations['gaze_duration'].std()
        ]
    })
    overall_stats.to_excel(writer, sheet_name='Overall Statistics', index=False)
    
    # Sheet 4: Session Summary
    # Get date range from session names (assuming format game_session_YYYY-MM-DD...)
    session_dates = []
    for session in df_durations['session_id'].unique():
        try:
            # Extract date from session name
            date_part = session.split('_')[2:5]  # game_session_YYYY-MM-DD
            if len(date_part) >= 3:
                date_str = '_'.join(date_part[:3])
                session_dates.append(pd.to_datetime(date_str))
        except:
            pass
    
    date_range = f"{min(session_dates).strftime('%Y-%m-%d')} to {max(session_dates).strftime('%Y-%m-%d')}" if session_dates else "N/A"
    
    session_summary = pd.DataFrame({
        'Metric': ['Total Sessions Scanned', 'Valid Sessions', 'Incomplete Sessions', 'Date Range'],
        'Value': [len(session_folders), valid_sessions_count, len(incomplete_sessions), date_range]
    })
    session_summary.to_excel(writer, sheet_name='Session Summary', index=False)
    
    # Sheet 5: Incomplete Sessions
    df_incomplete = pd.DataFrame(incomplete_sessions)
    if not df_incomplete.empty:
        df_incomplete.to_excel(writer, sheet_name='Incomplete Sessions', index=False)
    else:
        pd.DataFrame({'session_id': ['No incomplete sessions'], 'reason': ['']}).to_excel(
            writer, sheet_name='Incomplete Sessions', index=False)
    
    # Sheet 6: Gaze Point Stats
    df_gaze_stats = pd.DataFrame(gaze_point_stats)
    if not df_gaze_stats.empty:
        df_gaze_stats = df_gaze_stats.sort_values('session_id').reset_index(drop=True)
        df_gaze_stats.to_excel(writer, sheet_name='Gaze Point Stats', index=False)
    else:
        pd.DataFrame({'session_id': ['No data'], 'total_gaze_points': [0]}).to_excel(
            writer, sheet_name='Gaze Point Stats', index=False)
    
    # Sheet 7: Session Metadata
    df_metadata = pd.DataFrame(session_metadata_list)
    if not df_metadata.empty:
        df_metadata = df_metadata.sort_values('session_id').reset_index(drop=True)
        df_metadata.to_excel(writer, sheet_name='Session Metadata', index=False)
    else:
        pd.DataFrame({'session_id': ['No metadata available']}).to_excel(
            writer, sheet_name='Session Metadata', index=False)

# Apply formatting
wb = load_workbook(excel_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Bold headers and light blue background
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    
    # Freeze top row
    ws.freeze_panes = ws['A2']
    
    # Auto-fit columns
    for i, column in enumerate(ws.columns, start=1):
        max_length = 0
        column_letter = get_column_letter(i)
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long values
        ws.column_dimensions[column_letter].width = adjusted_width

wb.save(excel_path)

print(f"Excel file saved: {excel_path}")
print(f"  - Contains {len(wb.sheetnames)} sheets with formatted data")

# ==============================================================================
# ANALYSIS & VISUALIZATION
# ==============================================================================

# 1. Compute Gaze Statistics
print("\n--- Gaze Duration Statistics (per Question) ---")
stats = df_durations['gaze_duration'].agg(['min', 'max', 'mean', 'std'])
print(stats)

# Also stats per question ID
print("\n--- Average Duration per Question ID ---")
print(df_durations.groupby('question_id')['gaze_duration'].mean())

# Set style
sns.set_theme(style="whitegrid")

# CHART 1: Gaze Duration Distribution (Boxplot)
plt.figure(figsize=(10, 6))
sns.boxplot(x='question_id', y='gaze_duration', hue='question_id', data=df_durations, palette="viridis", legend=False)
plt.title('Distribution of Gaze Duration per Question', fontsize=15)
plt.xlabel('Question ID', fontsize=12)
plt.ylabel('Duration (seconds)', fontsize=12)
plt.tight_layout()
plt.savefig(os.path.join(OUTPUT_DIR, 'gaze_duration_boxplot.png'))
print(f"Chart saved: {os.path.join(OUTPUT_DIR, 'gaze_duration_boxplot.png')}")
plt.close()

# CHART 2: Gaze Heatmap
# Using all collected gaze points
if all_gaze_x and all_gaze_y:
    plt.figure(figsize=(12, 7))
    
    # Create a 2D histogram or KDE plot
    # Note: KDE can be slow with huge datasets. Histplot is faster.
    # We invert Y axis because screen coordinates usually have (0,0) at top-left
    # but plots usually have (0,0) at bottom-left.
    
    # Setup the plot limits based on screen resolution
    plt.xlim(0, screen_resolution[0])
    plt.ylim(screen_resolution[1], 0) # Inverted Y for screen coordinates
    
    # Plot background (optional, could be a screenshot of the game)
    # sns.kdeplot(x=all_gaze_x, y=all_gaze_y, fill=True, cmap="inferno", alpha=0.7, levels=20, thresh=0.05)
    
    # Using histplot for better performance with large data and "heatmap" look
    # Increased bins for higher resolution (e.g., 192x108 bins for ~10px blocks on 1080p)
    plt.hist2d(all_gaze_x, all_gaze_y, bins=[192, 108], range=[[0, screen_resolution[0]], [0, screen_resolution[1]]], cmap='inferno')
    plt.colorbar(label='Gaze Frequency')
    
    plt.title(f'Aggregate Gaze Heatmap (All Sessions, n={valid_sessions_count})', fontsize=15)
    plt.xlabel('Screen X (pixels)', fontsize=12)
    plt.ylabel('Screen Y (pixels)', fontsize=12)
    
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, 'gaze_heatmap.png'))
    print(f"Chart saved: {os.path.join(OUTPUT_DIR, 'gaze_heatmap.png')}")
    plt.close()
else:
    print("No gaze points found for heatmap.")

# CHART 3: Gaze Trajectory Heatmap (Relative to Screen Size)
# Uses normalized coordinates aggregated across all sessions
if all_gaze_u and all_gaze_v:
    plt.figure(figsize=(12, 7))
    # Plot normalized heatmap with higher bin resolution for smoother look
    plt.hist2d(all_gaze_u, all_gaze_v, bins=[200, 200], range=[[0, 1], [0, 1]], cmap='inferno')
    plt.colorbar(label='Normalized Gaze Frequency')
    # Invert Y to match screen coordinates (origin at top-left)
    plt.xlim(0, 1)
    plt.ylim(1, 0)
    plt.title(f'Aggregate Gaze Trajectory Heatmap (Relative, n={valid_sessions_count})', fontsize=15)
    plt.xlabel('Relative Screen X (0–1)', fontsize=12)
    plt.ylabel('Relative Screen Y (0–1)', fontsize=12)
    plt.tight_layout()
    plt.savefig(os.path.join(OUTPUT_DIR, 'gaze_trajectory_heatmap.png'))
    print(f"Chart saved: {os.path.join(OUTPUT_DIR, 'gaze_trajectory_heatmap.png')}")
    plt.close()
else:
    print("No normalized gaze points found for trajectory heatmap.")

print("\nAnalysis Complete.")

